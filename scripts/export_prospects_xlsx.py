#!/usr/bin/env python3
"""Build an Excel workbook from the private prospect list.

Four sheets, because a single dump of 288 rows is a file you open once:

  Summary          where the list actually is, by tier, region and wedge
  Outreach tracker the 151 tier 1-2 targets with empty status columns, so the
                   workbook is somewhere to work rather than something to read
  All targets      every row and every field, filterable
  Sources          the URLs behind the sourced rows, one per line

Written to the vault, not the repo: the source file is gitignored precisely so
this list stays private, and an xlsx of it belongs in the same place.

Run: python3 scripts/export_prospects_xlsx.py
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "dashboard" / "prospects.json"
OUT = Path.home() / "Documents" / "obsidian" / "Beverage-AI Radar" / "_private"

INK = "1B1B17"
ACCENT = "2D56CD"
TIER_FILL = {1: "E8F0E4", 2: "E6ECF7", 3: "F5F0E2", 4: "F1ECF5", 5: "F2F2F0"}
TIER_LABEL = {1: "Best fit, will pay", 2: "Corporate, long cycle", 3: "Volume play",
              4: "Channel multiplier", 5: "Adjacent buyer"}

THIN = Side(style="thin", color="D8D8D4")
BORDER = Border(bottom=THIN)


def header(ws, row: int, labels: list[str], widths: list[int], freeze: bool = True) -> None:
    for i, (label, w) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.row_dimensions[row].height = 26


def write_rows(ws, start: int, rows: list[list], tiers: list[int] | None = None) -> None:
    for r, values in enumerate(rows, start=start):
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=v)
            if v is None:
                c.value = None
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=10)
            c.border = BORDER
            if tiers:
                c.fill = PatternFill("solid", fgColor=TIER_FILL.get(tiers[r - start], "FFFFFF"))


def summary_sheet(wb, rows):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Prospects"
    ws["A1"].font = Font(bold=True, size=16, color=INK)
    ws["A2"] = f"{len(rows)} targets · generated {date.today().isoformat()}"
    ws["A3"] = "PRIVATE. Mirrors a gitignored file; never published. Do not share."
    ws["A3"].font = Font(italic=True, color="A03030", size=10)

    r = 5
    for title, key in (("By tier", "tier"), ("By region", "region"),
                       ("By wedge", "wedge_group"), ("By vertical", "vertical")):
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11)
        r += 1
        header(ws, r, [title.replace("By ", "").title(), "Count", "Share"], [34, 10, 10], freeze=False)
        r += 1
        tally = Counter(x.get(key) for x in rows if x.get(key) is not None)
        for k, n in sorted(tally.items(), key=lambda kv: -kv[1]):
            label = f"{k} — {TIER_LABEL[k]}" if key == "tier" else str(k)
            ws.cell(row=r, column=1, value=label).font = Font(size=10)
            ws.cell(row=r, column=2, value=n).font = Font(size=10)
            cell = ws.cell(row=r, column=3, value=n / len(rows))
            cell.number_format = "0%"
            cell.font = Font(size=10)
            r += 1
        r += 1
    ws.sheet_view.showGridLines = False


def tracker_sheet(wb, rows):
    """Tier 1-2 only: the rows you would actually email, with room to record it."""
    ws = wb.create_sheet("Outreach tracker")
    # Domain check sits before the empty columns on purpose: a row whose site
    # no longer resolves is one you want to see BEFORE you write the email,
    # not after.
    cols = ["Tier", "Company", "Region", "Segment", "Wedge", "Who to approach",
            "Domain check", "Status", "Date contacted", "Response", "Next step", "Notes"]
    widths = [6, 34, 18, 22, 24, 40, 30, 14, 14, 16, 24, 34]
    ws["A1"] = "Tier 1-2 targets. The last five columns are yours to fill in."
    ws["A1"].font = Font(italic=True, size=10, color="666560")
    header(ws, 2, cols, widths)
    live = [x for x in rows if x["tier"] <= 2]
    live.sort(key=lambda x: (x["tier"], x["region"], x["company"]))
    data = [[x["tier"], x["company"], x["region"], x.get("segment", ""),
             x.get("wedge_group") or x.get("wedge", ""), x.get("entry", ""),
             x.get("domain_status", ""),
             None, None, None, None, None] for x in live]
    write_rows(ws, 3, data, [x["tier"] for x in live])
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{2 + len(data)}"
    ws.sheet_view.showGridLines = False
    return len(data)


def all_sheet(wb, rows):
    ws = wb.create_sheet("All targets")
    cols = ["Tier", "Company", "Region", "Vertical", "HQ", "Segment", "Pain",
            "Wedge", "Wedge group", "Who to approach", "Capabilities",
            "Sources", "Provenance", "Domain check"]
    widths = [6, 32, 16, 11, 24, 22, 62, 30, 20, 38, 26, 9, 14, 30]
    header(ws, 1, cols, widths)
    rows = sorted(rows, key=lambda x: (x["tier"], x["region"], x["company"]))
    data = [[x["tier"], x["company"], x["region"], x.get("vertical", ""), x.get("hq", ""),
             x.get("segment", ""), x.get("pain", ""), x.get("wedge", ""),
             x.get("wedge_group", ""), x.get("entry", ""),
             ", ".join(x.get("capabilities") or []),
             len(x.get("source_urls") or []), x.get("discovered_by", "original"),
             x.get("domain_status", "")]
            for x in rows]
    write_rows(ws, 2, data, [x["tier"] for x in rows])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{1 + len(data)}"
    ws.sheet_view.showGridLines = False


def sources_sheet(wb, rows):
    ws = wb.create_sheet("Sources")
    header(ws, 1, ["Company", "Region", "Tier", "Source URL"], [34, 16, 6, 96])
    data = [[x["company"], x["region"], x["tier"], u]
            for x in sorted(rows, key=lambda r: (r["tier"], r["company"]))
            for u in (x.get("source_urls") or [])]
    write_rows(ws, 2, data)
    for r in range(2, 2 + len(data)):
        ws.cell(row=r, column=4).font = Font(size=9, color=ACCENT, underline="single")
    ws.auto_filter.ref = f"A1:D{1 + len(data)}"
    ws.sheet_view.showGridLines = False
    return len(data)


def main() -> int:
    if not SRC.exists():
        print(f"no prospect list at {SRC}")
        return 1
    rows = json.loads(SRC.read_text())
    wb = Workbook()
    wb.remove(wb.active)
    summary_sheet(wb, rows)
    n_track = tracker_sheet(wb, rows)
    all_sheet(wb, rows)
    n_src = sources_sheet(wb, rows)
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / "Prospects.xlsx"
    wb.save(path)
    print(f"{len(rows)} targets -> {path}")
    print(f"  Summary · Outreach tracker ({n_track} rows) · All targets ({len(rows)}) · Sources ({n_src})")
    print(f"  {path.stat().st_size / 1024:.0f} kb")
    return 0


if __name__ == "__main__":
    sys.exit(main())
